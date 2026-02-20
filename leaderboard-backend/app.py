"""
Multi-Tenant Embeddable Leaderboard Platform — Backend API
FastAPI + SQLite backend for multi-tenant fishing tournament leaderboards.
"""

import io
import os
import re
import json
import sqlite3
import secrets
from datetime import datetime, timedelta
from typing import Optional, List
from contextlib import contextmanager

from fastapi import FastAPI, HTTPException, Query, Depends, UploadFile, File, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse, StreamingResponse, RedirectResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel
from jose import jwt, JWTError
from passlib.context import CryptContext
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.utils import get_column_letter

# ============================================================================
# Configuration
# ============================================================================

_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
DATABASE_PATH = os.environ.get("DATABASE_PATH", os.path.join(_THIS_DIR, "leaderboard.db"))
JWT_SECRET = os.environ.get("JWT_SECRET", secrets.token_hex(32))
JWT_ALGORITHM = "HS256"
JWT_EXPIRE_HOURS = 24
UPLOAD_DIR = os.environ.get("UPLOAD_DIR", os.path.join(_THIS_DIR, "uploads"))
PARENT_DIR = os.path.dirname(_THIS_DIR)
DEFAULT_ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "steve.poland@cynoscionenvironmental.com")
DEFAULT_ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "fishleader4416")

os.makedirs(os.path.dirname(DATABASE_PATH), exist_ok=True)
os.makedirs(UPLOAD_DIR, exist_ok=True)

app = FastAPI(
    title="Leaderboard API",
    description="Multi-tenant fishing tournament leaderboard backend",
    version="1.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.middleware("http")
async def embed_headers_middleware(request: Request, call_next):
    response = await call_next(request)
    path = request.url.path
    if path.startswith("/embed/"):
        response.headers["X-Frame-Options"] = "ALLOWALL"
        response.headers["Content-Security-Policy"] = "frame-ancestors *"
    else:
        response.headers["X-Frame-Options"] = "SAMEORIGIN"
    return response


pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer(auto_error=False)

# ============================================================================
# Database Setup
# ============================================================================

def init_database():
    conn = sqlite3.connect(DATABASE_PATH)
    c = conn.cursor()

    c.execute("""
        CREATE TABLE IF NOT EXISTS series (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            slug TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL,
            year INTEGER,
            description TEXT,
            total_events INTEGER,
            best_of INTEGER,
            participation_points REAL DEFAULT 50,
            logo_path TEXT,
            status TEXT DEFAULT 'active',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            series_id INTEGER NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT DEFAULT 'admin',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (series_id) REFERENCES series(id) ON DELETE CASCADE
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS tournaments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            series_id INTEGER NOT NULL,
            event_name TEXT NOT NULL,
            event_number INTEGER,
            event_date TEXT,
            status TEXT DEFAULT 'upcoming',
            FOREIGN KEY (series_id) REFERENCES series(id) ON DELETE CASCADE
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            series_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            category_group TEXT,
            scoring_type TEXT DEFAULT 'points',
            is_standalone INTEGER DEFAULT 0,
            applies_to TEXT,
            unit TEXT DEFAULT 'pts',
            sort_order INTEGER DEFAULT 0,
            FOREIGN KEY (series_id) REFERENCES series(id) ON DELETE CASCADE
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS participants (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            series_id INTEGER NOT NULL,
            boat_name TEXT,
            captain TEXT,
            owner TEXT,
            angler_name TEXT,
            participant_type TEXT DEFAULT 'boat',
            boat_type TEXT,
            boat_id INTEGER,
            sonar INTEGER DEFAULT 0,
            homeport TEXT,
            photo TEXT,
            website TEXT,
            FOREIGN KEY (series_id) REFERENCES series(id) ON DELETE CASCADE,
            FOREIGN KEY (boat_id) REFERENCES participants(id) ON DELETE SET NULL
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS points (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            series_id INTEGER NOT NULL,
            tournament_id INTEGER NOT NULL,
            participant_id INTEGER NOT NULL,
            category_id INTEGER NOT NULL,
            points REAL NOT NULL,
            notes TEXT,
            UNIQUE(tournament_id, participant_id, category_id),
            FOREIGN KEY (series_id) REFERENCES series(id) ON DELETE CASCADE,
            FOREIGN KEY (tournament_id) REFERENCES tournaments(id) ON DELETE CASCADE,
            FOREIGN KEY (participant_id) REFERENCES participants(id) ON DELETE CASCADE,
            FOREIGN KEY (category_id) REFERENCES categories(id) ON DELETE CASCADE
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS scoring_rules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            series_id INTEGER NOT NULL,
            label TEXT NOT NULL,
            value TEXT NOT NULL,
            is_penalty INTEGER DEFAULT 0,
            sort_order INTEGER DEFAULT 0,
            FOREIGN KEY (series_id) REFERENCES series(id) ON DELETE CASCADE
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS tournament_participation (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            series_id INTEGER NOT NULL,
            tournament_id INTEGER NOT NULL,
            participant_id INTEGER NOT NULL,
            participated INTEGER DEFAULT 0,
            release_time TEXT,
            verified INTEGER DEFAULT 0,
            UNIQUE(tournament_id, participant_id),
            FOREIGN KEY (series_id) REFERENCES series(id) ON DELETE CASCADE,
            FOREIGN KEY (tournament_id) REFERENCES tournaments(id) ON DELETE CASCADE,
            FOREIGN KEY (participant_id) REFERENCES participants(id) ON DELETE CASCADE
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS weight_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            series_id INTEGER NOT NULL,
            tournament_id INTEGER NOT NULL,
            participant_id INTEGER NOT NULL,
            category_id INTEGER NOT NULL,
            weight REAL NOT NULL,
            angler_name TEXT,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (series_id) REFERENCES series(id) ON DELETE CASCADE,
            FOREIGN KEY (tournament_id) REFERENCES tournaments(id) ON DELETE CASCADE,
            FOREIGN KEY (participant_id) REFERENCES participants(id) ON DELETE CASCADE,
            FOREIGN KEY (category_id) REFERENCES categories(id) ON DELETE CASCADE
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS leaderboard_panels (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            series_id INTEGER NOT NULL,
            title TEXT NOT NULL,
            panel_type TEXT NOT NULL DEFAULT 'custom',
            content_json TEXT,
            visible INTEGER DEFAULT 1,
            sort_order INTEGER DEFAULT 0,
            FOREIGN KEY (series_id) REFERENCES series(id) ON DELETE CASCADE
        )
    """)

    # Migration: add columns to existing databases
    for col_sql in [
        "ALTER TABLE series ADD COLUMN primary_color TEXT DEFAULT '#0e8a7d'",
        "ALTER TABLE series ADD COLUMN accent_color TEXT DEFAULT '#b07d3a'",
        "ALTER TABLE series ADD COLUMN is_single_tournament INTEGER DEFAULT 0",
        "ALTER TABLE participants ADD COLUMN email TEXT",
        "ALTER TABLE categories ADD COLUMN include_in_overall INTEGER DEFAULT NULL",
        "ALTER TABLE categories ADD COLUMN points_per_fish REAL DEFAULT NULL",
        "ALTER TABLE categories ADD COLUMN is_aggregate INTEGER DEFAULT 0",
        "ALTER TABLE points ADD COLUMN fish_count INTEGER DEFAULT NULL",
    ]:
        try:
            c.execute(col_sql)
        except sqlite3.OperationalError:
            pass  # Column already exists

    conn.commit()
    conn.close()


@contextmanager
def get_db():
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        yield conn
    finally:
        conn.close()


# ============================================================================
# Pydantic Models
# ============================================================================

class SeriesUpdate(BaseModel):
    name: Optional[str] = None
    year: Optional[int] = None
    description: Optional[str] = None
    total_events: Optional[int] = None
    best_of: Optional[int] = None
    participation_points: Optional[float] = None
    logo_path: Optional[str] = None
    status: Optional[str] = None
    primary_color: Optional[str] = None
    accent_color: Optional[str] = None
    is_single_tournament: Optional[bool] = None

class SeriesCreate(BaseModel):
    slug: str
    name: str
    year: Optional[int] = None
    description: Optional[str] = None
    total_events: Optional[int] = None
    best_of: Optional[int] = 3
    participation_points: Optional[float] = 50
    logo_path: Optional[str] = None
    primary_color: Optional[str] = "#0e8a7d"
    accent_color: Optional[str] = "#b07d3a"
    is_single_tournament: Optional[bool] = False

class AdminUserCreate(BaseModel):
    email: str
    password: str
    series_slug: str
    role: Optional[str] = "admin"

class TournamentCreate(BaseModel):
    event_name: str
    event_number: Optional[int] = None
    event_date: Optional[str] = None
    status: Optional[str] = "upcoming"

class TournamentUpdate(BaseModel):
    event_name: Optional[str] = None
    event_number: Optional[int] = None
    event_date: Optional[str] = None
    status: Optional[str] = None

class CategoryCreate(BaseModel):
    name: str
    category_group: Optional[str] = None
    scoring_type: Optional[str] = "points"
    is_standalone: Optional[bool] = False
    applies_to: Optional[str] = None
    unit: Optional[str] = "pts"
    sort_order: Optional[int] = 0
    include_in_overall: Optional[bool] = None
    points_per_fish: Optional[float] = None
    is_aggregate: Optional[bool] = False

class ScoringRuleCreate(BaseModel):
    label: str
    value: str
    is_penalty: Optional[bool] = False
    sort_order: Optional[int] = 0

class ScoringRuleUpdate(BaseModel):
    label: Optional[str] = None
    value: Optional[str] = None
    is_penalty: Optional[bool] = None
    sort_order: Optional[int] = None

class ParticipantCreate(BaseModel):
    boat_name: Optional[str] = None
    captain: Optional[str] = None
    owner: Optional[str] = None
    angler_name: Optional[str] = None
    participant_type: Optional[str] = "boat"
    boat_type: Optional[str] = None
    boat_id: Optional[int] = None
    sonar: Optional[bool] = False
    homeport: Optional[str] = None
    photo: Optional[str] = None
    website: Optional[str] = None
    email: Optional[str] = None

class ParticipantUpdate(BaseModel):
    boat_name: Optional[str] = None
    captain: Optional[str] = None
    owner: Optional[str] = None
    angler_name: Optional[str] = None
    participant_type: Optional[str] = None
    boat_type: Optional[str] = None
    boat_id: Optional[int] = None
    sonar: Optional[bool] = None
    homeport: Optional[str] = None
    photo: Optional[str] = None
    website: Optional[str] = None
    email: Optional[str] = None

class PointEntry(BaseModel):
    tournament_id: int
    participant_id: int
    category_id: int
    points: float
    notes: Optional[str] = None
    fish_count: Optional[int] = None

class PointsBatch(BaseModel):
    points: List[PointEntry]

class TournamentParticipationEntry(BaseModel):
    tournament_id: int
    participant_id: int
    participated: Optional[bool] = None
    release_time: Optional[str] = None
    verified: Optional[bool] = None

class TournamentParticipationBatch(BaseModel):
    entries: List[TournamentParticipationEntry]

class WeightEntryCreate(BaseModel):
    tournament_id: int
    participant_id: int
    category_id: int
    weight: float
    angler_name: Optional[str] = None
    notes: Optional[str] = None

class LeaderboardPanelCreate(BaseModel):
    title: str
    panel_type: Optional[str] = "custom"
    content_json: Optional[str] = None
    visible: Optional[bool] = True
    sort_order: Optional[int] = 0

class LeaderboardPanelUpdate(BaseModel):
    title: Optional[str] = None
    content_json: Optional[str] = None
    visible: Optional[bool] = None
    sort_order: Optional[int] = None

class LoginRequest(BaseModel):
    email: str
    password: str

class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str


# ============================================================================
# Leaderboard Panel Seeding
# ============================================================================

def seed_default_panels(conn, series_id):
    """Create default Scoring Reference and Release Totals panels if none exist for this series."""
    count = conn.execute(
        "SELECT COUNT(*) as cnt FROM leaderboard_panels WHERE series_id = ?", (series_id,)
    ).fetchone()["cnt"]
    if count == 0:
        conn.execute(
            "INSERT INTO leaderboard_panels (series_id, title, panel_type, visible, sort_order) VALUES (?, ?, ?, ?, ?)",
            (series_id, "Scoring Reference", "scoring_reference", 1, 0),
        )
        conn.execute(
            "INSERT INTO leaderboard_panels (series_id, title, panel_type, visible, sort_order) VALUES (?, ?, ?, ?, ?)",
            (series_id, "Series Release Totals", "release_totals", 1, 1),
        )
        conn.commit()


# ============================================================================
# Auth Helpers
# ============================================================================

def create_token(user_id: int, series_id: int, role: str) -> str:
    expire = datetime.utcnow() + timedelta(hours=JWT_EXPIRE_HOURS)
    return jwt.encode(
        {"sub": str(user_id), "series_id": series_id, "role": role, "exp": expire},
        JWT_SECRET,
        algorithm=JWT_ALGORITHM,
    )


def decode_token(token: str) -> dict:
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid or expired token")


def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    if credentials is None:
        raise HTTPException(status_code=401, detail="Not authenticated")
    return decode_token(credentials.credentials)


def require_series_access(slug: str, user: dict = Depends(get_current_user)):
    """Verify user has access to the given series."""
    with get_db() as conn:
        series = conn.execute("SELECT id FROM series WHERE slug = ?", (slug,)).fetchone()
        if not series:
            raise HTTPException(status_code=404, detail="Series not found")
        if user["role"] != "super_admin" and user["series_id"] != series["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
    return user


def require_super_admin(user: dict = Depends(get_current_user)):
    """Only allow super_admin users."""
    if user.get("role") != "super_admin":
        raise HTTPException(status_code=403, detail="Super admin access required")
    return user


# ============================================================================
# Helper: resolve series slug to id
# ============================================================================

def get_series_by_slug(slug: str, conn) -> dict:
    row = conn.execute("SELECT * FROM series WHERE slug = ?", (slug,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Series not found")
    return dict(row)


# ============================================================================
# Standings Computation (Python port of JS computeStandings)
# ============================================================================

def compute_standings(
    conn,
    series_id: int,
    category_filter: str = "overall",
    group_filter: str = "all",
    search_term: str = "",
):
    series = dict(conn.execute("SELECT * FROM series WHERE id = ?", (series_id,)).fetchone())
    best_of = series["best_of"] or 3
    part_pts = series["participation_points"] or 0

    tournaments = [
        dict(r) for r in conn.execute(
            "SELECT * FROM tournaments WHERE series_id = ? ORDER BY event_number", (series_id,)
        ).fetchall()
    ]
    tourn_ids = [t["id"] for t in tournaments]

    if series.get("is_single_tournament"):
        best_of = len(tourn_ids) or 999
        part_pts = 0

    categories = [
        dict(r) for r in conn.execute(
            "SELECT * FROM categories WHERE series_id = ? ORDER BY sort_order", (series_id,)
        ).fetchall()
    ]

    all_participants = [
        dict(r) for r in conn.execute(
            "SELECT * FROM participants WHERE series_id = ?", (series_id,)
        ).fetchall()
    ]

    all_points = [
        dict(r) for r in conn.execute(
            "SELECT * FROM points WHERE series_id = ?", (series_id,)
        ).fetchall()
    ]

    all_tp = [
        dict(r) for r in conn.execute(
            "SELECT * FROM tournament_participation WHERE series_id = ?", (series_id,)
        ).fetchall()
    ]
    has_any_tp = len(all_tp) > 0
    # Build lookup: (tournament_id, participant_id) -> tp record
    tp_lookup = {}
    for tp in all_tp:
        tp_lookup[(tp["tournament_id"], tp["participant_id"])] = tp

    # Helpers
    boats = [p for p in all_participants if p["participant_type"] == "boat"]
    def get_individuals(ptype):
        return [p for p in all_participants if p["participant_type"] == ptype]
    def get_cat_by_id(cid):
        return next((c for c in categories if c["id"] == cid), None)
    def get_boat_by_id(bid):
        return next((p for p in boats if p["id"] == bid), None)

    is_individual_view = False
    use_participation = False
    unit = "pts"
    selected_cat = None

    def is_overall_cat(c):
        """Check if a category counts toward overall standings."""
        inc = c.get("include_in_overall")
        if inc == 1:
            return True
        if inc is not None and inc != 1:
            return False
        return not c["is_standalone"]

    if category_filter == "overall":
        candidates = list(boats)
        relevant_cat_ids = [c["id"] for c in categories if is_overall_cat(c)]
        use_participation = True
    else:
        try:
            cat_id = int(category_filter)
        except ValueError:
            cat_id = None
        selected_cat = get_cat_by_id(cat_id) if cat_id else None
        if not selected_cat:
            return {
                "standings": [], "tournaments": tournaments,
                "isIndividualView": False, "unit": "pts",
                "useParticipation": False, "selectedCat": None,
            }

        unit = selected_cat.get("unit", "pts")

        if selected_cat.get("applies_to") in ("sonar", "non_sonar"):
            want_sonar = 1 if selected_cat["applies_to"] == "sonar" else 0
            candidates = [b for b in boats if b["sonar"] == want_sonar]
            relevant_cat_ids = [c["id"] for c in categories if is_overall_cat(c)]
            use_participation = True
        elif selected_cat.get("applies_to"):
            candidates = get_individuals(selected_cat["applies_to"])
            is_individual_view = True
            relevant_cat_ids = [c["id"] for c in categories if is_overall_cat(c)]
            use_participation = True
        else:
            candidates = list(boats)
            relevant_cat_ids = [selected_cat["id"]]
            use_participation = selected_cat["scoring_type"] != "weight"

    # Group filter (boats only)
    if not is_individual_view:
        if group_filter == "private":
            candidates = [p for p in candidates if p.get("boat_type") == "private"]
        elif group_filter == "charter":
            candidates = [p for p in candidates if p.get("boat_type") == "charter"]
        elif group_filter == "sonar":
            candidates = [p for p in candidates if p.get("sonar")]
        elif group_filter == "non_sonar":
            candidates = [p for p in candidates if not p.get("sonar")]

    # Search
    if search_term:
        q = search_term.lower()
        filtered = []
        for p in candidates:
            if q in (p.get("boat_name") or "").lower():
                filtered.append(p); continue
            if q in (p.get("captain") or "").lower():
                filtered.append(p); continue
            if q in (p.get("angler_name") or "").lower():
                filtered.append(p); continue
            if p.get("boat_id"):
                b = get_boat_by_id(p["boat_id"])
                if b and q in (b.get("boat_name") or "").lower():
                    filtered.append(p); continue
        candidates = filtered

    # Build standings
    is_single = bool(series.get("is_single_tournament"))
    standings = []
    for part in candidates:
        row = {
            "participant": part,
            "tournamentPoints": {},
            "totalAll": 0,
            "bestOfScore": 0,
            "tournamentsEntered": 0,
            "counted": {},
            "participationBonus": 0,
            "tournamentVerified": {},
            "tournamentFishCounts": {},
        }
        if is_single:
            row["categoryPoints"] = {}

        latest_release_time = None

        for tid in tourn_ids:
            pts = [
                p for p in all_points
                if p["participant_id"] == part["id"]
                and p["tournament_id"] == tid
                and p["category_id"] in relevant_cat_ids
            ]
            total = sum(p["points"] for p in pts)
            row["tournamentPoints"][str(tid)] = total

            # Tournament participation: explicit checkbox OR has scores
            tp = tp_lookup.get((tid, part["id"]))
            if tp and tp["participated"]:
                row["tournamentsEntered"] += 1
            elif total > 0:
                # Auto-count as participated if they have any scores recorded
                row["tournamentsEntered"] += 1

            # Track verification status
            if tp:
                row["tournamentVerified"][str(tid)] = bool(tp["verified"])
                if tp.get("release_time"):
                    latest_release_time = max(latest_release_time or "", tp["release_time"])

            row["totalAll"] += total

            # Per-category fish count breakdowns
            for p in pts:
                if p.get("fish_count") is not None:
                    key = f"{tid}_{p['category_id']}"
                    row["tournamentFishCounts"][key] = p["fish_count"]

            # Per-category breakdowns for single-tournament mode
            if is_single:
                for p in pts:
                    cid = str(p["category_id"])
                    row["categoryPoints"][cid] = row["categoryPoints"].get(cid, 0) + p["points"]

        row["latestReleaseTime"] = latest_release_time or "99:99"

        # Best-of-N
        scored = [
            {"tid": tid, "pts": row["tournamentPoints"].get(str(tid), 0)}
            for tid in tourn_ids
        ]
        scored = [s for s in scored if s["pts"] > 0]
        scored.sort(key=lambda s: s["pts"], reverse=True)

        top_n = scored[:best_of]
        best_of_sum = sum(s["pts"] for s in top_n)
        row["participationBonus"] = row["tournamentsEntered"] * part_pts if use_participation else 0
        row["bestOfScore"] = best_of_sum + row["participationBonus"]

        for item in top_n:
            row["counted"][str(item["tid"])] = True

        standings.append(row)

    # Sort with release time tiebreaker (earlier release time wins)
    standings.sort(key=lambda r: (
        -r["bestOfScore"],
        -r["totalAll"],
        r.get("latestReleaseTime", "99:99"),
    ))

    # Assign ranks with tie handling (include release time in tie check)
    for i, s in enumerate(standings):
        if (i > 0
            and s["bestOfScore"] == standings[i-1]["bestOfScore"]
            and s["totalAll"] == standings[i-1]["totalAll"]
            and s.get("latestReleaseTime", "99:99") == standings[i-1].get("latestReleaseTime", "99:99")):
            s["rank"] = standings[i-1]["rank"]
        else:
            s["rank"] = i + 1

    result = {
        "standings": standings,
        "tournaments": tournaments,
        "isIndividualView": is_individual_view,
        "unit": unit,
        "useParticipation": use_participation,
        "selectedCat": selected_cat,
    }
    if is_single:
        result["scoringCategories"] = [
            c for c in categories if c["id"] in relevant_cat_ids
        ]
    return result


# ============================================================================
# Auth Endpoints
# ============================================================================

@app.post("/api/auth/login")
def login(req: LoginRequest):
    with get_db() as conn:
        user = conn.execute("SELECT * FROM users WHERE email = ?", (req.email,)).fetchone()
        if not user or not pwd_context.verify(req.password, user["password_hash"]):
            raise HTTPException(status_code=401, detail="Invalid credentials")
        token = create_token(user["id"], user["series_id"], user["role"])
        return {"token": token, "email": user["email"], "role": user["role"]}


@app.get("/api/auth/me")
def auth_me(user: dict = Depends(get_current_user)):
    with get_db() as conn:
        row = conn.execute(
            "SELECT u.id, u.email, u.role, u.series_id, s.slug as series_slug, s.name as series_name "
            "FROM users u JOIN series s ON u.series_id = s.id WHERE u.id = ?",
            (int(user["sub"]),)
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="User not found")
        return dict(row)


@app.post("/api/auth/change-password")
def change_password(req: ChangePasswordRequest, user: dict = Depends(get_current_user)):
    if len(req.new_password) < 8:
        raise HTTPException(status_code=400, detail="New password must be at least 8 characters")
    with get_db() as conn:
        row = conn.execute("SELECT * FROM users WHERE id = ?", (int(user["sub"]),)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="User not found")
        if not pwd_context.verify(req.current_password, row["password_hash"]):
            raise HTTPException(status_code=401, detail="Current password is incorrect")
        new_hash = pwd_context.hash(req.new_password)
        conn.execute("UPDATE users SET password_hash = ? WHERE id = ?", (new_hash, row["id"]))
        conn.commit()
        return {"success": True}


# ============================================================================
# Super-Admin Endpoints (must be before /api/{slug} catch-all)
# ============================================================================

@app.get("/api/admin/series")
def admin_list_series(user: dict = Depends(require_super_admin)):
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM series ORDER BY name").fetchall()
        return [dict(r) for r in rows]


@app.post("/api/admin/series")
def admin_create_series(data: SeriesCreate, user: dict = Depends(require_super_admin)):
    if not re.match(r'^[a-z0-9][a-z0-9_-]{1,48}[a-z0-9]$', data.slug):
        raise HTTPException(status_code=400, detail="Slug must be 3-50 chars: lowercase letters, numbers, hyphens, underscores")
    with get_db() as conn:
        existing = conn.execute("SELECT id FROM series WHERE slug = ?", (data.slug,)).fetchone()
        if existing:
            raise HTTPException(status_code=409, detail="Series slug already exists")
        conn.execute(
            "INSERT INTO series (slug, name, year, description, total_events, best_of, participation_points, logo_path, primary_color, accent_color, is_single_tournament, status) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'active')",
            (data.slug, data.name, data.year, data.description, data.total_events,
             data.best_of, data.participation_points, data.logo_path, data.primary_color, data.accent_color,
             1 if data.is_single_tournament else 0),
        )
        conn.commit()
        new_series = conn.execute("SELECT * FROM series WHERE slug = ?", (data.slug,)).fetchone()
        seed_default_panels(conn, new_series["id"])
        return dict(new_series)


@app.delete("/api/admin/series/{slug}")
def admin_delete_series(slug: str, user: dict = Depends(require_super_admin)):
    with get_db() as conn:
        series = conn.execute("SELECT id FROM series WHERE slug = ?", (slug,)).fetchone()
        if not series:
            raise HTTPException(status_code=404, detail="Series not found")
        sid = series["id"]
        conn.execute("DELETE FROM weight_entries WHERE series_id = ?", (sid,))
        conn.execute("DELETE FROM tournament_participation WHERE series_id = ?", (sid,))
        conn.execute("DELETE FROM points WHERE series_id = ?", (sid,))
        conn.execute("DELETE FROM participants WHERE series_id = ?", (sid,))
        conn.execute("DELETE FROM categories WHERE series_id = ?", (sid,))
        conn.execute("DELETE FROM leaderboard_panels WHERE series_id = ?", (sid,))
        conn.execute("DELETE FROM scoring_rules WHERE series_id = ?", (sid,))
        conn.execute("DELETE FROM tournaments WHERE series_id = ?", (sid,))
        conn.execute("DELETE FROM users WHERE series_id = ?", (sid,))
        conn.execute("DELETE FROM series WHERE id = ?", (sid,))
        conn.commit()
        return {"deleted": True}


@app.post("/api/admin/upload-logo")
async def admin_upload_logo(file: UploadFile = File(...), user: dict = Depends(require_super_admin)):
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in (".jpg", ".jpeg", ".png", ".gif", ".webp"):
        raise HTTPException(status_code=400, detail="Only image files allowed")
    safe_name = f"logo_{secrets.token_hex(8)}{ext}"
    path = os.path.join(UPLOAD_DIR, safe_name)
    content = await file.read()
    with open(path, "wb") as f:
        f.write(content)
    return {"filename": safe_name}


@app.get("/api/admin/users")
def admin_list_users(user: dict = Depends(require_super_admin)):
    with get_db() as conn:
        rows = conn.execute(
            "SELECT u.id, u.email, u.role, u.series_id, s.slug as series_slug, s.name as series_name "
            "FROM users u JOIN series s ON u.series_id = s.id ORDER BY u.email"
        ).fetchall()
        return [dict(r) for r in rows]


@app.post("/api/admin/users")
def admin_create_user(data: AdminUserCreate, user: dict = Depends(require_super_admin)):
    with get_db() as conn:
        series = conn.execute("SELECT id FROM series WHERE slug = ?", (data.series_slug,)).fetchone()
        if not series:
            raise HTTPException(status_code=404, detail="Series not found")
        existing = conn.execute("SELECT id FROM users WHERE email = ?", (data.email,)).fetchone()
        if existing:
            raise HTTPException(status_code=409, detail="User with this email already exists")
        hashed = pwd_context.hash(data.password)
        conn.execute(
            "INSERT INTO users (series_id, email, password_hash, role) VALUES (?, ?, ?, ?)",
            (series["id"], data.email, hashed, data.role),
        )
        conn.commit()
        return {"created": True, "email": data.email, "series_slug": data.series_slug}


# ============================================================================
# Public Endpoints
# ============================================================================

@app.get("/api/{slug}")
def get_series(slug: str):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        return s


@app.get("/api/{slug}/standings")
def get_standings(
    slug: str,
    category: str = Query("overall"),
    group: str = Query("all"),
    search: str = Query(""),
):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        return compute_standings(conn, s["id"], category, group, search)


@app.get("/api/{slug}/tournaments")
def list_tournaments(slug: str):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        rows = conn.execute(
            "SELECT * FROM tournaments WHERE series_id = ? ORDER BY event_number", (s["id"],)
        ).fetchall()
        return [dict(r) for r in rows]


@app.get("/api/{slug}/categories")
def list_categories(slug: str):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        rows = conn.execute(
            "SELECT * FROM categories WHERE series_id = ? ORDER BY sort_order", (s["id"],)
        ).fetchall()
        return [dict(r) for r in rows]


@app.post("/api/{slug}/categories")
def create_category(slug: str, data: CategoryCreate, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        if user["role"] != "super_admin" and user["series_id"] != s["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        inc_overall = None
        if data.include_in_overall is not None:
            inc_overall = 1 if data.include_in_overall else 0
        cur = conn.execute(
            "INSERT INTO categories (series_id, name, category_group, scoring_type, is_standalone, applies_to, unit, sort_order, include_in_overall, points_per_fish, is_aggregate) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (s["id"], data.name, data.category_group, data.scoring_type,
             1 if data.is_standalone else 0, data.applies_to, data.unit, data.sort_order,
             inc_overall, data.points_per_fish, 1 if data.is_aggregate else 0),
        )
        conn.commit()
        return dict(conn.execute("SELECT * FROM categories WHERE id = ?", (cur.lastrowid,)).fetchone())


@app.delete("/api/{slug}/categories/{cid}")
def delete_category(slug: str, cid: int, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        if user["role"] != "super_admin" and user["series_id"] != s["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        affected = conn.execute(
            "DELETE FROM categories WHERE id = ? AND series_id = ?", (cid, s["id"])
        ).rowcount
        conn.commit()
        if not affected:
            raise HTTPException(status_code=404, detail="Category not found")
        return {"deleted": True}


@app.get("/api/{slug}/scoring-rules")
def list_scoring_rules(slug: str):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        rows = conn.execute(
            "SELECT * FROM scoring_rules WHERE series_id = ? ORDER BY sort_order", (s["id"],)
        ).fetchall()
        return [dict(r) for r in rows]


@app.post("/api/{slug}/scoring-rules")
def create_scoring_rule(slug: str, data: ScoringRuleCreate, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        if user["role"] != "super_admin" and user["series_id"] != s["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        cur = conn.execute(
            "INSERT INTO scoring_rules (series_id, label, value, is_penalty, sort_order) VALUES (?, ?, ?, ?, ?)",
            (s["id"], data.label, data.value, 1 if data.is_penalty else 0, data.sort_order),
        )
        conn.commit()
        return dict(conn.execute("SELECT * FROM scoring_rules WHERE id = ?", (cur.lastrowid,)).fetchone())


@app.put("/api/{slug}/scoring-rules/{rid}")
def update_scoring_rule(slug: str, rid: int, data: ScoringRuleUpdate, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        if user["role"] != "super_admin" and user["series_id"] != s["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        existing = conn.execute(
            "SELECT * FROM scoring_rules WHERE id = ? AND series_id = ?", (rid, s["id"])
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Scoring rule not found")
        updates = {}
        for k, v in data.model_dump().items():
            if v is not None:
                if k == "is_penalty":
                    updates[k] = 1 if v else 0
                else:
                    updates[k] = v
        if not updates:
            return dict(existing)
        set_clause = ", ".join(f"{k} = ?" for k in updates)
        vals = list(updates.values()) + [rid]
        conn.execute(f"UPDATE scoring_rules SET {set_clause} WHERE id = ?", vals)
        conn.commit()
        return dict(conn.execute("SELECT * FROM scoring_rules WHERE id = ?", (rid,)).fetchone())


@app.delete("/api/{slug}/scoring-rules/{rid}")
def delete_scoring_rule(slug: str, rid: int, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        if user["role"] != "super_admin" and user["series_id"] != s["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        affected = conn.execute(
            "DELETE FROM scoring_rules WHERE id = ? AND series_id = ?", (rid, s["id"])
        ).rowcount
        conn.commit()
        if not affected:
            raise HTTPException(status_code=404, detail="Scoring rule not found")
        return {"deleted": True}


@app.get("/api/{slug}/participants")
def list_participants(slug: str):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        rows = conn.execute(
            "SELECT * FROM participants WHERE series_id = ?", (s["id"],)
        ).fetchall()
        return [dict(r) for r in rows]


@app.get("/api/{slug}/participants/{participant_id}")
def get_participant(slug: str, participant_id: int):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        row = conn.execute(
            "SELECT * FROM participants WHERE id = ? AND series_id = ?",
            (participant_id, s["id"]),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Participant not found")
        return dict(row)


@app.get("/api/{slug}/points")
def list_points(slug: str):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        rows = conn.execute(
            "SELECT * FROM points WHERE series_id = ?", (s["id"],)
        ).fetchall()
        return [dict(r) for r in rows]


# ============================================================================
# Admin Endpoints — Series
# ============================================================================

@app.put("/api/{slug}/series")
def update_series(slug: str, data: SeriesUpdate, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        if user["role"] != "super_admin" and user["series_id"] != s["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        updates = {k: v for k, v in data.model_dump().items() if v is not None}
        if "is_single_tournament" in updates:
            updates["is_single_tournament"] = 1 if updates["is_single_tournament"] else 0
        if not updates:
            return dict(s)
        updates["updated_at"] = datetime.utcnow().isoformat()
        set_clause = ", ".join(f"{k} = ?" for k in updates)
        vals = list(updates.values()) + [s["id"]]
        conn.execute(f"UPDATE series SET {set_clause} WHERE id = ?", vals)
        conn.commit()
        return dict(conn.execute("SELECT * FROM series WHERE id = ?", (s["id"],)).fetchone())


# ============================================================================
# Admin Endpoints — Tournaments
# ============================================================================

@app.post("/api/{slug}/tournaments")
def create_tournament(slug: str, data: TournamentCreate, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        if user["role"] != "super_admin" and user["series_id"] != s["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        cur = conn.execute(
            "INSERT INTO tournaments (series_id, event_name, event_number, event_date, status) VALUES (?, ?, ?, ?, ?)",
            (s["id"], data.event_name, data.event_number, data.event_date, data.status),
        )
        conn.commit()
        return dict(conn.execute("SELECT * FROM tournaments WHERE id = ?", (cur.lastrowid,)).fetchone())


@app.put("/api/{slug}/tournaments/{tid}")
def update_tournament(slug: str, tid: int, data: TournamentUpdate, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        if user["role"] != "super_admin" and user["series_id"] != s["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        existing = conn.execute(
            "SELECT * FROM tournaments WHERE id = ? AND series_id = ?", (tid, s["id"])
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Tournament not found")
        updates = {k: v for k, v in data.model_dump().items() if v is not None}
        if not updates:
            return dict(existing)
        set_clause = ", ".join(f"{k} = ?" for k in updates)
        vals = list(updates.values()) + [tid]
        conn.execute(f"UPDATE tournaments SET {set_clause} WHERE id = ?", vals)
        conn.commit()
        return dict(conn.execute("SELECT * FROM tournaments WHERE id = ?", (tid,)).fetchone())


@app.delete("/api/{slug}/tournaments/{tid}")
def delete_tournament(slug: str, tid: int, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        if user["role"] != "super_admin" and user["series_id"] != s["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        affected = conn.execute(
            "DELETE FROM tournaments WHERE id = ? AND series_id = ?", (tid, s["id"])
        ).rowcount
        conn.commit()
        if not affected:
            raise HTTPException(status_code=404, detail="Tournament not found")
        return {"deleted": True}


# ============================================================================
# Admin Endpoints — Participants
# ============================================================================

@app.post("/api/{slug}/participants")
def create_participant(slug: str, data: ParticipantCreate, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        if user["role"] != "super_admin" and user["series_id"] != s["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        cur = conn.execute(
            """INSERT INTO participants
               (series_id, boat_name, captain, owner, angler_name, participant_type,
                boat_type, boat_id, sonar, homeport, photo, website, email)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (s["id"], data.boat_name, data.captain, data.owner, data.angler_name,
             data.participant_type, data.boat_type, data.boat_id,
             1 if data.sonar else 0, data.homeport, data.photo, data.website, data.email),
        )
        conn.commit()
        return dict(conn.execute("SELECT * FROM participants WHERE id = ?", (cur.lastrowid,)).fetchone())


@app.put("/api/{slug}/participants/{pid}")
def update_participant(slug: str, pid: int, data: ParticipantUpdate, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        if user["role"] != "super_admin" and user["series_id"] != s["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        existing = conn.execute(
            "SELECT * FROM participants WHERE id = ? AND series_id = ?", (pid, s["id"])
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Participant not found")
        updates = {}
        for k, v in data.model_dump().items():
            if v is not None:
                if k == "sonar":
                    updates[k] = 1 if v else 0
                else:
                    updates[k] = v
        if not updates:
            return dict(existing)
        set_clause = ", ".join(f"{k} = ?" for k in updates)
        vals = list(updates.values()) + [pid]
        conn.execute(f"UPDATE participants SET {set_clause} WHERE id = ?", vals)
        conn.commit()
        return dict(conn.execute("SELECT * FROM participants WHERE id = ?", (pid,)).fetchone())


@app.delete("/api/{slug}/participants/{pid}")
def delete_participant(slug: str, pid: int, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        if user["role"] != "super_admin" and user["series_id"] != s["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        affected = conn.execute(
            "DELETE FROM participants WHERE id = ? AND series_id = ?", (pid, s["id"])
        ).rowcount
        conn.commit()
        if not affected:
            raise HTTPException(status_code=404, detail="Participant not found")
        return {"deleted": True}


# ============================================================================
# Admin Endpoints — Points
# ============================================================================

@app.post("/api/{slug}/points")
def upsert_points(slug: str, data: PointsBatch, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        if user["role"] != "super_admin" and user["series_id"] != s["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        # Build category lookup for fish_count auto-calc
        cat_lookup = {}
        cat_rows = conn.execute("SELECT id, points_per_fish FROM categories WHERE series_id = ?", (s["id"],)).fetchall()
        for cr in cat_rows:
            cat_lookup[cr["id"]] = dict(cr)

        inserted = 0
        updated = 0
        deleted = 0
        for entry in data.points:
            actual_points = entry.points
            fish_count = entry.fish_count

            # Auto-calculate points from fish_count if category has points_per_fish
            cat_info = cat_lookup.get(entry.category_id)
            if fish_count is not None and cat_info and cat_info.get("points_per_fish"):
                actual_points = fish_count * cat_info["points_per_fish"]

            existing = conn.execute(
                "SELECT id FROM points WHERE tournament_id = ? AND participant_id = ? AND category_id = ?",
                (entry.tournament_id, entry.participant_id, entry.category_id),
            ).fetchone()
            if actual_points <= 0 and (fish_count is None or fish_count <= 0):
                if existing:
                    conn.execute("DELETE FROM points WHERE id = ?", (existing["id"],))
                    deleted += 1
            elif existing:
                conn.execute(
                    "UPDATE points SET points = ?, notes = ?, fish_count = ? WHERE id = ?",
                    (actual_points, entry.notes, fish_count, existing["id"]),
                )
                updated += 1
            else:
                conn.execute(
                    "INSERT INTO points (series_id, tournament_id, participant_id, category_id, points, notes, fish_count) VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (s["id"], entry.tournament_id, entry.participant_id, entry.category_id, actual_points, entry.notes, fish_count),
                )
                inserted += 1
        conn.commit()
        return {"inserted": inserted, "updated": updated, "deleted": deleted}


@app.delete("/api/{slug}/points/{point_id}")
def delete_point(slug: str, point_id: int, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        if user["role"] != "super_admin" and user["series_id"] != s["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        affected = conn.execute(
            "DELETE FROM points WHERE id = ? AND series_id = ?", (point_id, s["id"])
        ).rowcount
        conn.commit()
        if not affected:
            raise HTTPException(status_code=404, detail="Point entry not found")
        return {"deleted": True}


# ============================================================================
# Tournament Participation Endpoints
# ============================================================================

@app.get("/api/{slug}/tournament-participation")
def list_tournament_participation(slug: str):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        rows = conn.execute(
            "SELECT * FROM tournament_participation WHERE series_id = ?", (s["id"],)
        ).fetchall()
        return [dict(r) for r in rows]


@app.put("/api/{slug}/tournament-participation/{tid}/{pid}")
def upsert_tournament_participation(
    slug: str, tid: int, pid: int,
    data: TournamentParticipationEntry,
    user: dict = Depends(get_current_user),
):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        if user["role"] != "super_admin" and user["series_id"] != s["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        existing = conn.execute(
            "SELECT * FROM tournament_participation WHERE tournament_id = ? AND participant_id = ?",
            (tid, pid),
        ).fetchone()
        if existing:
            updates = {}
            if data.participated is not None:
                updates["participated"] = 1 if data.participated else 0
            if data.release_time is not None:
                updates["release_time"] = data.release_time
            if data.verified is not None:
                updates["verified"] = 1 if data.verified else 0
            if updates:
                set_clause = ", ".join(f"{k} = ?" for k in updates)
                vals = list(updates.values()) + [existing["id"]]
                conn.execute(f"UPDATE tournament_participation SET {set_clause} WHERE id = ?", vals)
        else:
            conn.execute(
                "INSERT INTO tournament_participation (series_id, tournament_id, participant_id, participated, release_time, verified) VALUES (?, ?, ?, ?, ?, ?)",
                (s["id"], tid, pid,
                 1 if data.participated else 0,
                 data.release_time,
                 1 if data.verified else 0),
            )
        conn.commit()
        row = conn.execute(
            "SELECT * FROM tournament_participation WHERE tournament_id = ? AND participant_id = ?",
            (tid, pid),
        ).fetchone()
        return dict(row)


@app.post("/api/{slug}/tournament-participation/batch")
def batch_tournament_participation(
    slug: str, data: TournamentParticipationBatch,
    user: dict = Depends(get_current_user),
):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        if user["role"] != "super_admin" and user["series_id"] != s["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        upserted = 0
        for entry in data.entries:
            existing = conn.execute(
                "SELECT * FROM tournament_participation WHERE tournament_id = ? AND participant_id = ?",
                (entry.tournament_id, entry.participant_id),
            ).fetchone()
            if existing:
                updates = {}
                if entry.participated is not None:
                    updates["participated"] = 1 if entry.participated else 0
                if entry.release_time is not None:
                    updates["release_time"] = entry.release_time
                if entry.verified is not None:
                    updates["verified"] = 1 if entry.verified else 0
                if updates:
                    set_clause = ", ".join(f"{k} = ?" for k in updates)
                    vals = list(updates.values()) + [existing["id"]]
                    conn.execute(f"UPDATE tournament_participation SET {set_clause} WHERE id = ?", vals)
                    upserted += 1
            else:
                conn.execute(
                    "INSERT INTO tournament_participation (series_id, tournament_id, participant_id, participated, release_time, verified) VALUES (?, ?, ?, ?, ?, ?)",
                    (s["id"], entry.tournament_id, entry.participant_id,
                     1 if entry.participated else 0,
                     entry.release_time,
                     1 if entry.verified else 0),
                )
                upserted += 1
        conn.commit()
        return {"upserted": upserted}


@app.post("/api/{slug}/tournament-participation/verify/{tid}")
def verify_tournament(slug: str, tid: int, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        if user["role"] != "super_admin" and user["series_id"] != s["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        conn.execute(
            "UPDATE tournament_participation SET verified = 1 WHERE series_id = ? AND tournament_id = ?",
            (s["id"], tid),
        )
        conn.commit()
        return {"verified": True, "tournament_id": tid}


# ============================================================================
# Weight Entries Endpoints
# ============================================================================

@app.post("/api/{slug}/weight-entries")
def add_weight_entry(slug: str, data: WeightEntryCreate, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        if user["role"] != "super_admin" and user["series_id"] != s["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        conn.execute(
            "INSERT INTO weight_entries (series_id, tournament_id, participant_id, category_id, weight, angler_name, notes) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (s["id"], data.tournament_id, data.participant_id, data.category_id,
             data.weight, data.angler_name, data.notes),
        )
        # Update points table with MAX weight
        max_row = conn.execute(
            "SELECT MAX(weight) as max_w FROM weight_entries WHERE series_id = ? AND tournament_id = ? AND participant_id = ? AND category_id = ?",
            (s["id"], data.tournament_id, data.participant_id, data.category_id),
        ).fetchone()
        max_weight = max_row["max_w"] if max_row else data.weight
        existing_pt = conn.execute(
            "SELECT id FROM points WHERE tournament_id = ? AND participant_id = ? AND category_id = ?",
            (data.tournament_id, data.participant_id, data.category_id),
        ).fetchone()
        if existing_pt:
            conn.execute("UPDATE points SET points = ? WHERE id = ?", (max_weight, existing_pt["id"]))
        else:
            conn.execute(
                "INSERT INTO points (series_id, tournament_id, participant_id, category_id, points) VALUES (?, ?, ?, ?, ?)",
                (s["id"], data.tournament_id, data.participant_id, data.category_id, max_weight),
            )
        conn.commit()
        return {"max_weight": max_weight}


@app.get("/api/{slug}/weight-entries")
def list_weight_entries(
    slug: str,
    tournament_id: Optional[int] = None,
    participant_id: Optional[int] = None,
    category_id: Optional[int] = None,
):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        query = "SELECT * FROM weight_entries WHERE series_id = ?"
        params = [s["id"]]
        if tournament_id:
            query += " AND tournament_id = ?"
            params.append(tournament_id)
        if participant_id:
            query += " AND participant_id = ?"
            params.append(participant_id)
        if category_id:
            query += " AND category_id = ?"
            params.append(category_id)
        query += " ORDER BY created_at DESC"
        rows = conn.execute(query, params).fetchall()
        return [dict(r) for r in rows]


@app.delete("/api/{slug}/weight-entries/{entry_id}")
def delete_weight_entry(slug: str, entry_id: int, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        if user["role"] != "super_admin" and user["series_id"] != s["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        entry = conn.execute(
            "SELECT * FROM weight_entries WHERE id = ? AND series_id = ?", (entry_id, s["id"])
        ).fetchone()
        if not entry:
            raise HTTPException(status_code=404, detail="Weight entry not found")
        conn.execute("DELETE FROM weight_entries WHERE id = ?", (entry_id,))
        # Recalculate max for this combination
        max_row = conn.execute(
            "SELECT MAX(weight) as max_w FROM weight_entries WHERE series_id = ? AND tournament_id = ? AND participant_id = ? AND category_id = ?",
            (s["id"], entry["tournament_id"], entry["participant_id"], entry["category_id"]),
        ).fetchone()
        if max_row and max_row["max_w"] is not None:
            conn.execute(
                "UPDATE points SET points = ? WHERE tournament_id = ? AND participant_id = ? AND category_id = ?",
                (max_row["max_w"], entry["tournament_id"], entry["participant_id"], entry["category_id"]),
            )
        else:
            conn.execute(
                "DELETE FROM points WHERE tournament_id = ? AND participant_id = ? AND category_id = ?",
                (entry["tournament_id"], entry["participant_id"], entry["category_id"]),
            )
        conn.commit()
        return {"deleted": True}


# ============================================================================
# Release Summary Endpoint
# ============================================================================

@app.get("/api/{slug}/release-summary/{tournament_id}")
def release_summary(slug: str, tournament_id: int):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        categories = [dict(r) for r in conn.execute(
            "SELECT * FROM categories WHERE series_id = ? AND points_per_fish IS NOT NULL ORDER BY sort_order",
            (s["id"],),
        ).fetchall()]
        result = []
        for cat in categories:
            row = conn.execute(
                "SELECT SUM(fish_count) as total_fish, SUM(points) as total_points FROM points WHERE series_id = ? AND tournament_id = ? AND category_id = ? AND fish_count IS NOT NULL",
                (s["id"], tournament_id, cat["id"]),
            ).fetchone()
            result.append({
                "category": cat["name"],
                "category_id": cat["id"],
                "total_fish": row["total_fish"] or 0,
                "total_points": row["total_points"] or 0,
            })
        return result


def _get_release_totals(conn, series_id):
    """Helper: series-wide release totals across all tournaments, broken down by category and tournament."""
    categories = [dict(r) for r in conn.execute(
        "SELECT * FROM categories WHERE series_id = ? AND (points_per_fish IS NOT NULL OR category_group = 'Release') ORDER BY sort_order",
        (series_id,),
    ).fetchall()]
    tournaments = [dict(r) for r in conn.execute(
        "SELECT * FROM tournaments WHERE series_id = ? ORDER BY event_number", (series_id,)
    ).fetchall()]
    result = []
    for cat in categories:
        cat_totals = {"category": cat["name"], "category_id": cat["id"], "total_fish": 0, "total_points": 0, "by_tournament": []}
        for t in tournaments:
            row = conn.execute(
                "SELECT COUNT(*) as entry_count, SUM(fish_count) as fish, SUM(points) as pts FROM points WHERE series_id = ? AND tournament_id = ? AND category_id = ? AND points > 0",
                (series_id, t["id"], cat["id"]),
            ).fetchone()
            fish = row["fish"] or 0
            pts = row["pts"] or 0
            entries = row["entry_count"] or 0
            cat_totals["by_tournament"].append({
                "tournament_id": t["id"],
                "tournament_name": t["event_name"],
                "event_number": t["event_number"],
                "fish_count": fish,
                "points": pts,
                "entries": entries,
            })
            cat_totals["total_fish"] += fish
            cat_totals["total_points"] += pts
        result.append(cat_totals)
    return result


@app.get("/api/{slug}/release-totals")
def release_totals(slug: str):
    """Series-wide release totals across all tournaments, broken down by category and tournament."""
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        return _get_release_totals(conn, s["id"])


# ============================================================================
# Leaderboard Panels Endpoints
# ============================================================================

@app.get("/api/{slug}/leaderboard-panels")
def list_leaderboard_panels(slug: str):
    """Public: returns visible panels with resolved content."""
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        sid = s["id"]
        panels = [dict(r) for r in conn.execute(
            "SELECT * FROM leaderboard_panels WHERE series_id = ? AND visible = 1 ORDER BY sort_order",
            (sid,),
        ).fetchall()]
        result = []
        for p in panels:
            item = {"id": p["id"], "title": p["title"], "panel_type": p["panel_type"], "sort_order": p["sort_order"]}
            if p["panel_type"] == "scoring_reference":
                rules = [dict(r) for r in conn.execute(
                    "SELECT * FROM scoring_rules WHERE series_id = ? ORDER BY sort_order", (sid,)
                ).fetchall()]
                item["content"] = rules
            elif p["panel_type"] == "release_totals":
                item["content"] = _get_release_totals(conn, sid)
            else:
                try:
                    item["content"] = json.loads(p["content_json"]) if p["content_json"] else []
                except (json.JSONDecodeError, TypeError):
                    item["content"] = []
            result.append(item)
        return result


@app.get("/api/{slug}/leaderboard-panels/admin")
def list_leaderboard_panels_admin(slug: str, user: dict = Depends(get_current_user)):
    """Admin: returns ALL panels (including hidden), raw metadata."""
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        if user["role"] != "super_admin" and user["series_id"] != s["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        panels = [dict(r) for r in conn.execute(
            "SELECT * FROM leaderboard_panels WHERE series_id = ? ORDER BY sort_order",
            (s["id"],),
        ).fetchall()]
        return panels


@app.post("/api/{slug}/leaderboard-panels")
def create_leaderboard_panel(slug: str, data: LeaderboardPanelCreate, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        if user["role"] != "super_admin" and user["series_id"] != s["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        sid = s["id"]
        # Prevent duplicate built-in types
        if data.panel_type in ("scoring_reference", "release_totals"):
            existing = conn.execute(
                "SELECT id FROM leaderboard_panels WHERE series_id = ? AND panel_type = ?",
                (sid, data.panel_type),
            ).fetchone()
            if existing:
                raise HTTPException(status_code=409, detail=f"A {data.panel_type} panel already exists for this series")
        cur = conn.execute(
            "INSERT INTO leaderboard_panels (series_id, title, panel_type, content_json, visible, sort_order) VALUES (?, ?, ?, ?, ?, ?)",
            (sid, data.title, data.panel_type, data.content_json,
             1 if data.visible else 0, data.sort_order),
        )
        conn.commit()
        return dict(conn.execute("SELECT * FROM leaderboard_panels WHERE id = ?", (cur.lastrowid,)).fetchone())


@app.put("/api/{slug}/leaderboard-panels/{pid}")
def update_leaderboard_panel(slug: str, pid: int, data: LeaderboardPanelUpdate, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        if user["role"] != "super_admin" and user["series_id"] != s["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        existing = conn.execute(
            "SELECT * FROM leaderboard_panels WHERE id = ? AND series_id = ?", (pid, s["id"])
        ).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Panel not found")
        updates = {}
        for k, v in data.model_dump().items():
            if v is not None:
                if k == "visible":
                    updates[k] = 1 if v else 0
                else:
                    updates[k] = v
        if not updates:
            return dict(existing)
        set_clause = ", ".join(f"{k} = ?" for k in updates)
        vals = list(updates.values()) + [pid]
        conn.execute(f"UPDATE leaderboard_panels SET {set_clause} WHERE id = ?", vals)
        conn.commit()
        return dict(conn.execute("SELECT * FROM leaderboard_panels WHERE id = ?", (pid,)).fetchone())


@app.delete("/api/{slug}/leaderboard-panels/{pid}")
def delete_leaderboard_panel(slug: str, pid: int, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        if user["role"] != "super_admin" and user["series_id"] != s["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        affected = conn.execute(
            "DELETE FROM leaderboard_panels WHERE id = ? AND series_id = ?", (pid, s["id"])
        ).rowcount
        conn.commit()
        if not affected:
            raise HTTPException(status_code=404, detail="Panel not found")
        return {"deleted": True}


# ============================================================================
# Excel Template Download & Import
# ============================================================================

@app.get("/api/{slug}/template/{tournament_id}")
def download_template(slug: str, tournament_id: int, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        if user["role"] != "super_admin" and user["series_id"] != s["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        sid = s["id"]

        tourn = conn.execute(
            "SELECT * FROM tournaments WHERE id = ? AND series_id = ?", (tournament_id, sid)
        ).fetchone()
        if not tourn:
            raise HTTPException(status_code=404, detail="Tournament not found")

        categories = [dict(r) for r in conn.execute(
            "SELECT * FROM categories WHERE series_id = ? ORDER BY sort_order", (sid,)
        ).fetchall()]
        participants = [dict(r) for r in conn.execute(
            "SELECT * FROM participants WHERE series_id = ?", (sid,)
        ).fetchall()]
        existing_points = [dict(r) for r in conn.execute(
            "SELECT * FROM points WHERE series_id = ? AND tournament_id = ?", (sid, tournament_id)
        ).fetchall()]

        # Build lookup for existing points: (participant_id, category_id) -> points
        pts_lookup = {(p["participant_id"], p["category_id"]): p["points"] for p in existing_points}

        boats = [p for p in participants if p["participant_type"] == "boat"]
        lady_anglers = [p for p in participants if p["participant_type"] == "lady_angler"]
        junior_anglers = [p for p in participants if p["participant_type"] == "junior_angler"]
        junior_boys = [p for p in participants if p["participant_type"] == "junior_angler_boy"]
        junior_girls = [p for p in participants if p["participant_type"] == "junior_angler_girl"]

        # Boats: categories with no applies_to; individuals: their specific applies_to category
        boat_cats = [c for c in categories if not c.get("applies_to")]
        lady_cats = [c for c in categories if c.get("applies_to") == "lady_angler"]
        junior_cats = [c for c in categories if c.get("applies_to") == "junior_angler"]
        junior_boy_cats = [c for c in categories if c.get("applies_to") == "junior_angler_boy"]
        junior_girl_cats = [c for c in categories if c.get("applies_to") == "junior_angler_girl"]

        wb = Workbook()

        def build_sheet(ws, title, parts, cats, name_key, sub_key):
            ws.title = title
            if not parts or not cats:
                ws.append(["No data for this sheet"])
                return

            # Row 1: Headers
            headers = ["ID", name_key, sub_key] + [c["name"] for c in cats]
            ws.append(headers)

            # Row 2: Hidden meta row with category IDs
            meta = ["__meta__", "", ""] + [str(c["id"]) for c in cats]
            ws.append(meta)

            # Data rows
            for p in parts:
                name_val = p.get("boat_name") or p.get("angler_name") or ""
                if name_key == "Boat Name":
                    sub_val = p.get("captain") or ""
                else:
                    # Individual: show boat name
                    boat = next((b for b in boats if b["id"] == p.get("boat_id")), None)
                    sub_val = boat["boat_name"] if boat else ""
                row = [p["id"], name_val, sub_val]
                for c in cats:
                    val = pts_lookup.get((p["id"], c["id"]))
                    row.append(val if val is not None else None)
                ws.append(row)

            # Styling
            bold = Font(bold=True)
            for cell in ws[1]:
                cell.font = bold
                cell.alignment = Alignment(horizontal="center")

            # Hide meta row (row 2) and ID column (col A)
            ws.row_dimensions[2].hidden = True
            ws.column_dimensions["A"].hidden = True

            # Auto-size columns
            for col_idx in range(1, len(headers) + 1):
                max_len = max(
                    len(str(ws.cell(row=r, column=col_idx).value or ""))
                    for r in range(1, ws.max_row + 1)
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 30)

            # Number format for point cells (columns D onward, row 3+)
            for row in ws.iter_rows(min_row=3, min_col=4, max_col=len(headers)):
                for cell in row:
                    cell.number_format = numbers.FORMAT_NUMBER_00

        # First sheet (default) = Boats
        build_sheet(wb.active, "Boats", boats, boat_cats, "Boat Name", "Captain")

        # Lady Anglers sheet
        ws_lady = wb.create_sheet()
        build_sheet(ws_lady, "Lady Anglers", lady_anglers, lady_cats, "Angler Name", "Boat")

        # Junior Anglers sheet
        ws_junior = wb.create_sheet()
        build_sheet(ws_junior, "Junior Anglers", junior_anglers, junior_cats, "Angler Name", "Boat")

        # Junior Angler Boy sheet
        if junior_boys:
            ws_jb = wb.create_sheet()
            build_sheet(ws_jb, "Jr. Boys", junior_boys, junior_boy_cats or junior_cats, "Angler Name", "Boat")

        # Junior Angler Girl sheet
        if junior_girls:
            ws_jg = wb.create_sheet()
            build_sheet(ws_jg, "Jr. Girls", junior_girls, junior_girl_cats or junior_cats, "Angler Name", "Boat")

        # Write to bytes
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"Tournament_{tournament_id}_Results_Template.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )


@app.post("/api/{slug}/import-results/{tournament_id}")
async def import_results(slug: str, tournament_id: int, file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        if user["role"] != "super_admin" and user["series_id"] != s["id"]:
            raise HTTPException(status_code=403, detail="Access denied")
        sid = s["id"]

        tourn = conn.execute(
            "SELECT * FROM tournaments WHERE id = ? AND series_id = ?", (tournament_id, sid)
        ).fetchone()
        if not tourn:
            raise HTTPException(status_code=404, detail="Tournament not found")

        if not file.filename.endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="Only .xlsx files are accepted")

        content = await file.read()
        try:
            wb = load_workbook(io.BytesIO(content), data_only=True)
        except Exception:
            raise HTTPException(status_code=400, detail="Could not read Excel file")

        point_entries = []

        for ws in wb.worksheets:
            if ws.max_row < 3:
                continue

            # Row 2 is the meta row: col A = "__meta__", cols D+ = category IDs
            meta_marker = ws.cell(row=2, column=1).value
            if str(meta_marker).strip() != "__meta__":
                continue

            # Read category IDs from meta row (columns 4+)
            cat_ids = []
            for col_idx in range(4, ws.max_column + 1):
                val = ws.cell(row=2, column=col_idx).value
                try:
                    cat_ids.append(int(val))
                except (TypeError, ValueError):
                    cat_ids.append(None)

            # Data rows start at row 3
            for row_idx in range(3, ws.max_row + 1):
                pid_val = ws.cell(row=row_idx, column=1).value
                try:
                    participant_id = int(pid_val)
                except (TypeError, ValueError):
                    continue

                for i, cat_id in enumerate(cat_ids):
                    if cat_id is None:
                        continue
                    cell_val = ws.cell(row=row_idx, column=4 + i).value
                    try:
                        pts = float(cell_val) if cell_val is not None else 0
                    except (TypeError, ValueError):
                        pts = 0

                    point_entries.append(PointEntry(
                        tournament_id=tournament_id,
                        participant_id=participant_id,
                        category_id=cat_id,
                        points=pts,
                    ))

        # Upsert all points using the same logic as upsert_points
        inserted = 0
        updated = 0
        deleted = 0
        for entry in point_entries:
            existing = conn.execute(
                "SELECT id FROM points WHERE tournament_id = ? AND participant_id = ? AND category_id = ?",
                (entry.tournament_id, entry.participant_id, entry.category_id),
            ).fetchone()
            if entry.points <= 0:
                if existing:
                    conn.execute("DELETE FROM points WHERE id = ?", (existing["id"],))
                    deleted += 1
            elif existing:
                conn.execute(
                    "UPDATE points SET points = ?, notes = ? WHERE id = ?",
                    (entry.points, entry.notes, existing["id"]),
                )
                updated += 1
            else:
                conn.execute(
                    "INSERT INTO points (series_id, tournament_id, participant_id, category_id, points, notes) VALUES (?, ?, ?, ?, ?, ?)",
                    (sid, entry.tournament_id, entry.participant_id, entry.category_id, entry.points, entry.notes),
                )
                inserted += 1

        # Auto-mark tournament as completed if it was upcoming
        if dict(tourn)["status"] == "upcoming":
            conn.execute("UPDATE tournaments SET status = 'completed' WHERE id = ?", (tournament_id,))

        conn.commit()
        return {"inserted": inserted, "updated": updated, "deleted": deleted}


# ============================================================================
# Import / Export
# ============================================================================

@app.post("/api/{slug}/import")
async def import_data(slug: str, request: Request, user: dict = Depends(get_current_user)):
    """Import data. If JSON body provided, use it; otherwise read ncbs-data.json from disk."""
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        if user["role"] != "super_admin" and user["series_id"] != s["id"]:
            raise HTTPException(status_code=403, detail="Access denied")

        data = None
        body = await request.body()
        if body:
            try:
                data = json.loads(body)
            except json.JSONDecodeError:
                raise HTTPException(status_code=400, detail="Invalid JSON body")

        if not data:
            json_path = os.path.join(PARENT_DIR, "ncbs-data.json")
            if not os.path.exists(json_path):
                raise HTTPException(status_code=404, detail="ncbs-data.json not found in parent directory")
            with open(json_path, "r") as f:
                data = json.load(f)

        series_id = s["id"]

        # Update series fields
        sd = data.get("series", {})
        conn.execute(
            "UPDATE series SET name=?, year=?, description=?, total_events=?, best_of=?, participation_points=?, status=?, updated_at=? WHERE id=?",
            (sd.get("name", s["name"]), sd.get("year"), sd.get("description"),
             sd.get("total_events"), sd.get("best_of"), sd.get("participation_points", 50),
             sd.get("status", "active"), datetime.utcnow().isoformat(), series_id),
        )

        # Clear existing data for this series
        conn.execute("DELETE FROM points WHERE series_id = ?", (series_id,))
        conn.execute("DELETE FROM participants WHERE series_id = ?", (series_id,))
        conn.execute("DELETE FROM categories WHERE series_id = ?", (series_id,))
        conn.execute("DELETE FROM tournaments WHERE series_id = ?", (series_id,))

        # ID mapping: old JSON id -> new DB id
        tourn_map = {}
        cat_map = {}
        part_map = {}

        # Tournaments
        for t in data.get("tournaments", []):
            cur = conn.execute(
                "INSERT INTO tournaments (series_id, event_name, event_number, event_date, status) VALUES (?, ?, ?, ?, ?)",
                (series_id, t["event_name"], t.get("event_number"), t.get("event_date"), t.get("status", "upcoming")),
            )
            tourn_map[t["id"]] = cur.lastrowid

        # Categories
        for c in data.get("categories", []):
            inc_overall = c.get("include_in_overall")
            if inc_overall is not None:
                inc_overall = 1 if inc_overall else 0
            cur = conn.execute(
                "INSERT INTO categories (series_id, name, category_group, scoring_type, is_standalone, applies_to, unit, sort_order, include_in_overall, points_per_fish, is_aggregate) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (series_id, c["name"], c.get("category_group"), c.get("scoring_type", "points"),
                 1 if c.get("is_standalone") else 0, c.get("applies_to"), c.get("unit", "pts"), c.get("sort_order", 0),
                 inc_overall, c.get("points_per_fish"), 1 if c.get("is_aggregate") else 0),
            )
            cat_map[c["id"]] = cur.lastrowid

        # Participants — insert boats first (no boat_id), then individuals
        boats_data = [p for p in data.get("participants", []) if p.get("participant_type", "boat") == "boat"]
        individuals_data = [p for p in data.get("participants", []) if p.get("participant_type", "boat") != "boat"]

        for p in boats_data:
            cur = conn.execute(
                """INSERT INTO participants
                   (series_id, boat_name, captain, owner, angler_name, participant_type,
                    boat_type, boat_id, sonar, homeport, photo, website, email)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (series_id, p.get("boat_name"), p.get("captain"), p.get("owner"),
                 p.get("angler_name"), "boat", p.get("boat_type"), None,
                 1 if p.get("sonar") else 0, p.get("homeport"), p.get("photo"), p.get("website"), p.get("email")),
            )
            part_map[p["id"]] = cur.lastrowid

        for p in individuals_data:
            boat_ref = part_map.get(p.get("boat_id")) if p.get("boat_id") else None
            cur = conn.execute(
                """INSERT INTO participants
                   (series_id, boat_name, captain, owner, angler_name, participant_type,
                    boat_type, boat_id, sonar, homeport, photo, website, email)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (series_id, p.get("boat_name"), p.get("captain"), p.get("owner"),
                 p.get("angler_name"), p.get("participant_type", "boat"), p.get("boat_type"),
                 boat_ref, 1 if p.get("sonar") else 0, p.get("homeport"), p.get("photo"), p.get("website"), p.get("email")),
            )
            part_map[p["id"]] = cur.lastrowid

        # Points
        for pt in data.get("points", []):
            tid = tourn_map.get(pt["tournament_id"])
            pid = part_map.get(pt["participant_id"])
            cid = cat_map.get(pt["category_id"])
            if tid and pid and cid:
                conn.execute(
                    "INSERT INTO points (series_id, tournament_id, participant_id, category_id, points, notes, fish_count) VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (series_id, tid, pid, cid, pt["points"], pt.get("notes"), pt.get("fish_count")),
                )

        # Leaderboard Panels
        conn.execute("DELETE FROM leaderboard_panels WHERE series_id = ?", (series_id,))
        if data.get("leaderboard_panels"):
            for lp in data["leaderboard_panels"]:
                conn.execute(
                    "INSERT INTO leaderboard_panels (series_id, title, panel_type, content_json, visible, sort_order) VALUES (?, ?, ?, ?, ?, ?)",
                    (series_id, lp["title"], lp.get("panel_type", "custom"),
                     lp.get("content_json"), 1 if lp.get("visible", 1) else 0,
                     lp.get("sort_order", 0)),
                )
        else:
            # No panels in import data — seed defaults
            seed_default_panels(conn, series_id)

        conn.commit()
        return {
            "imported": True,
            "tournaments": len(tourn_map),
            "categories": len(cat_map),
            "participants": len(part_map),
            "points": len(data.get("points", [])),
        }


@app.post("/api/{slug}/export")
def export_data(slug: str, user: dict = Depends(get_current_user)):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        if user["role"] != "super_admin" and user["series_id"] != s["id"]:
            raise HTTPException(status_code=403, detail="Access denied")

        sid = s["id"]
        tournaments = [dict(r) for r in conn.execute("SELECT * FROM tournaments WHERE series_id = ? ORDER BY event_number", (sid,)).fetchall()]
        categories = [dict(r) for r in conn.execute("SELECT * FROM categories WHERE series_id = ? ORDER BY sort_order", (sid,)).fetchall()]
        participants = [dict(r) for r in conn.execute("SELECT * FROM participants WHERE series_id = ?", (sid,)).fetchall()]
        points = [dict(r) for r in conn.execute("SELECT * FROM points WHERE series_id = ?", (sid,)).fetchall()]
        leaderboard_panels = [dict(r) for r in conn.execute("SELECT * FROM leaderboard_panels WHERE series_id = ? ORDER BY sort_order", (sid,)).fetchall()]

        return {
            "series": {k: s[k] for k in ("name", "year", "description", "total_events", "best_of", "participation_points", "status")},
            "tournaments": tournaments,
            "categories": categories,
            "participants": participants,
            "points": points,
            "leaderboard_panels": leaderboard_panels,
        }


# ============================================================================
# Photo Upload
# ============================================================================

@app.post("/api/{slug}/upload")
async def upload_photo(slug: str, file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    with get_db() as conn:
        s = get_series_by_slug(slug, conn)
        if user["role"] != "super_admin" and user["series_id"] != s["id"]:
            raise HTTPException(status_code=403, detail="Access denied")

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in (".jpg", ".jpeg", ".png", ".gif", ".webp"):
        raise HTTPException(status_code=400, detail="Only image files allowed")

    safe_name = f"{secrets.token_hex(8)}{ext}"
    path = os.path.join(UPLOAD_DIR, safe_name)
    content = await file.read()
    with open(path, "wb") as f:
        f.write(content)
    return {"filename": safe_name}


# ============================================================================
# Startup: Init DB, seed default series + admin
# ============================================================================

SEED_SERIES_SLUG = os.environ.get("SEED_SERIES_SLUG", "ncbillfish")
SEED_SERIES_NAME = os.environ.get("SEED_SERIES_NAME", "NC Billfish Series")


@app.on_event("startup")
def startup():
    init_database()

    with get_db() as conn:
        # Only seed if series table is empty (first run)
        any_series = conn.execute("SELECT COUNT(*) as cnt FROM series").fetchone()
        if any_series["cnt"] == 0:
            conn.execute(
                "INSERT INTO series (slug, name, year, description, total_events, best_of, participation_points, logo_path, status) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (SEED_SERIES_SLUG, SEED_SERIES_NAME, 2026,
                 "Fishing tournament series",
                 8, 3, 50, "/static/NCBillfishSeries.jpg", "active"),
            )
            # Seed default scoring rules for the NC Billfish series
            seeded_series = conn.execute("SELECT id FROM series WHERE slug = ?", (SEED_SERIES_SLUG,)).fetchone()
            if seeded_series:
                default_rules = [
                    (seeded_series["id"], "Per tournament series fished", "50 pts", 0, 0),
                    (seeded_series["id"], "Per blue marlin released", "400 pts", 0, 1),
                    (seeded_series["id"], "Per white marlin, sailfish, or spearfish released", "125 pts", 0, 2),
                    (seeded_series["id"], "Per unverified blue marlin released", "125 pts", 0, 3),
                    (seeded_series["id"], 'Per blue marlin landed (min 110" or 400 lbs)', "1 pt/lb", 0, 4),
                    (seeded_series["id"], "Penalty for undersized fish", "-200 pts", 1, 5),
                ]
                conn.executemany(
                    "INSERT INTO scoring_rules (series_id, label, value, is_penalty, sort_order) VALUES (?, ?, ?, ?, ?)",
                    default_rules,
                )
            conn.commit()
            print(f"Seeded series: {SEED_SERIES_SLUG}")

        # Seed default panels for all series that have none
        all_series = conn.execute("SELECT id FROM series").fetchall()
        for s_row in all_series:
            seed_default_panels(conn, s_row["id"])

        # Create default admin if none exists
        existing_user = conn.execute("SELECT id FROM users WHERE email = ?", (DEFAULT_ADMIN_EMAIL,)).fetchone()
        if not existing_user:
            series = conn.execute("SELECT id FROM series ORDER BY id LIMIT 1").fetchone()
            if series:
                hashed = pwd_context.hash(DEFAULT_ADMIN_PASSWORD)
                conn.execute(
                    "INSERT INTO users (series_id, email, password_hash, role) VALUES (?, ?, ?, ?)",
                    (series["id"], DEFAULT_ADMIN_EMAIL, hashed, "super_admin"),
                )
                conn.commit()
                print(f"Default admin created: {DEFAULT_ADMIN_EMAIL}")


# ============================================================================
# Admin Landing Page
# ============================================================================

@app.get("/admin")
def serve_admin_landing():
    html_path = os.path.join(PARENT_DIR, "admin.html")
    if os.path.exists(html_path):
        return FileResponse(html_path, media_type="text/html")
    return HTMLResponse("<h1>Admin portal not found</h1>")


# ============================================================================
# Tenant Frontend Routes
# ============================================================================

@app.get("/t/{slug}")
def serve_tenant(slug: str):
    """Serve leaderboard.html for a specific series (validates slug exists)."""
    with get_db() as conn:
        series = conn.execute("SELECT id FROM series WHERE slug = ?", (slug,)).fetchone()
        if not series:
            raise HTTPException(status_code=404, detail="Series not found")
    html_path = os.path.join(PARENT_DIR, "leaderboard.html")
    if os.path.exists(html_path):
        return FileResponse(html_path, media_type="text/html")
    return HTMLResponse("<h1>Leaderboard not found</h1>")


@app.get("/embed/{slug}")
def serve_embed(slug: str):
    """Serve leaderboard.html for iframe embedding."""
    with get_db() as conn:
        series = conn.execute("SELECT id FROM series WHERE slug = ?", (slug,)).fetchone()
        if not series:
            raise HTTPException(status_code=404, detail="Series not found")
    html_path = os.path.join(PARENT_DIR, "leaderboard.html")
    if os.path.exists(html_path):
        return FileResponse(html_path, media_type="text/html")
    return HTMLResponse("<h1>Leaderboard not found</h1>")


@app.get("/widget/{slug}/leaderboard.js")
def serve_widget_js(slug: str, request: Request):
    """Serve a JS snippet that creates an auto-resizing iframe embed."""
    base_url = str(request.base_url).rstrip("/")
    js = f"""(function(){{
  var container = document.currentScript.getAttribute('data-container');
  var el = container ? document.getElementById(container) : document.currentScript.parentElement;
  if (!el) return;
  var iframe = document.createElement('iframe');
  iframe.src = '{base_url}/embed/{slug}';
  iframe.style.width = '100%';
  iframe.style.border = 'none';
  iframe.style.minHeight = '600px';
  iframe.setAttribute('scrolling', 'no');
  el.appendChild(iframe);
  window.addEventListener('message', function(e) {{
    if (e.data && e.data.type === 'leaderboard-resize' && e.data.slug === '{slug}') {{
      iframe.style.height = e.data.height + 'px';
    }}
  }});
}})();"""
    from starlette.responses import Response
    return Response(content=js, media_type="application/javascript")


@app.get("/")
def serve_root():
    """If 1 series, redirect to /t/{slug}. If multiple, show directory."""
    with get_db() as conn:
        rows = conn.execute("SELECT slug, name FROM series WHERE status = 'active' ORDER BY name").fetchall()
        if len(rows) == 0:
            return HTMLResponse("<h1>Leaderboard Platform</h1><p>No series configured. Visit <a href='/docs'>/docs</a> for the API.</p>")
        if len(rows) == 1:
            return RedirectResponse(url=f"/t/{rows[0]['slug']}", status_code=302)
        # Multiple series — show directory page
        items = "".join(
            f'<li style="margin:0.5rem 0;"><a href="/t/{r["slug"]}" style="font-size:1.1rem;">{r["name"]}</a></li>'
            for r in rows
        )
        html = f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><title>Tournament Leaderboards</title>
        <style>body{{font-family:sans-serif;max-width:600px;margin:2rem auto;padding:0 1rem;}}
        a{{color:#0e8a7d;}} a:hover{{color:#b07d3a;}}</style></head>
        <body><h1>Tournament Leaderboards</h1><ul>{items}</ul></body></html>"""
        return HTMLResponse(html)


# ============================================================================
# Static File Serving
# ============================================================================

# Serve uploaded photos
app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")

# Serve frontend assets from parent directory (leaderboard.html, logos, etc.)
app.mount("/static", StaticFiles(directory=PARENT_DIR), name="static")
